from openpyxl import Workbook
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import random


def create_log_sheet(name, code, plate, sdate, edate, from_place, distance, price_per_km, total_amount_display,
                     driver_name, odometer_start):
    wb = Workbook()
    ws = wb.active
    cell_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    bold_font = Font(name='Calibri', bold=True)
    big_bold_font = Font(name='Calibri', bold=True, size=14)
    thin = Side(border_style="thin", color="000000")
    cell_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws.title = "VOE Logsheet 2023"

    # ws.merge_cells('A1:B1')
    ws['A1'] = "Vehicle Log Book"
    ws['A2'] = "Name"
    ws['B2'] = name

    ws['C2'] = "Employee Code"
    ws['D2'] = code

    ws['E2'] = 'Odometer reading (at start of period)'
    ws['F2'] = odometer_start

    ws['A3'] = 'Period'
    ws['B3'] = 'From: {}'.format(datetime.strftime(sdate, '%d-%m-%Y'))
    ws['C3'] = 'To: {}'.format(datetime.strftime(edate, '%d-%m-%Y'))

    ws['D3'] = 'Number Plate'
    ws['E3'] = plate

    ws['A4'] = 'Particulars'

    ws.merge_cells('B4:C4')
    ws['B5'] = 'Journey'

    ws.merge_cells('D4:F4')
    ws['B4'] = 'Odometer Reading'

    ws['G4'] = 'Reason for Trip'
    ws['H4'] = "Driver's Signature"
    ws['I4'] = 'Fuel Rate'

    ws['A5'] = 'Date'
    ws['B5'] = 'From'
    ws['C5'] = 'To'
    ws['D5'] = 'Start'
    ws['E5'] = 'Finish'
    ws['F5'] = 'Dist(in km)'

    ws['I5'] = "per km(INR {}/km))".format(str(price_per_km))
    i = 6
    total_dist = 0
    total_amount_display = 0
    while sdate <= edate:
        print(sdate)
        #print(edate)
        if sdate.weekday() in [5, 6]:
            print(sdate.weekday())
            sdate = sdate + timedelta(days=1)
            odometer_start = odometer_start + random.randint(100, 200)
            continue
        ws['A{}'.format(str(i))] = datetime.strftime(sdate, '%d-%m-%Y')
        ws['B{}'.format(str(i))] = from_place
        ws['C{}'.format(str(i))] = 'Gurugram'
        ws['D{}'.format(str(i))] = str(odometer_start)
        odometer_end = odometer_start + random.randint(57, 65)
        ws['E{}'.format(str(i))] = str(odometer_end)
        ws['F{}'.format(str(i))] = odometer_end - odometer_start
        ws['G{}'.format(str(i))] = 'Official'
        ws['H{}'.format(str(i))] = driver_name
        ws['I{}'.format(str(i))] = (odometer_end - odometer_start) * price_per_km
        total_dist += (odometer_end - odometer_start)
        total_amount_display += (odometer_end - odometer_start) * price_per_km
        odometer_start = odometer_end
        i += 1
        ws['A{}'.format(str(i))] = datetime.strftime(sdate, '%d-%m-%Y')
        ws['B{}'.format(str(i))] = 'Gurugram'
        ws['C{}'.format(str(i))] = from_place
        ws['D{}'.format(str(i))] = str(odometer_start)
        odometer_end = odometer_start + random.randint(57, 65)
        ws['E{}'.format(str(i))] = str(odometer_end)
        ws['F{}'.format(str(i))] = odometer_end - odometer_start
        ws['G{}'.format(str(i))] = 'Official'
        ws['H{}'.format(str(i))] = driver_name
        ws['I{}'.format(str(i))] = (odometer_end - odometer_start) * price_per_km
        total_amount_display += (odometer_end - odometer_start) * price_per_km
        total_dist += (odometer_end - odometer_start)
        odometer_start = odometer_end + random.randint(57, 65)
        i += 1
        sdate = sdate + timedelta(days=1)

    ws.merge_cells('A{}:E{}'.format(i, i))
    ws['A{}'.format(i)] = 'Total'
    ws['F{}'.format(i)] = str(total_dist)
    ws['I{}'.format(i)] = str(total_amount_display)
    print('total', total_amount_display)
    file_name = 'logsheet' + '_' + 'srishti' + '_' +datetime.today().strftime('%b') + '_' + '.xlsx'
    wb.save('/home/delhivery/Log/' + file_name)


if __name__ == '__main__':
    name = 'Srishti Sharma'
    code = 'SSN034206'
    plate = 'DL 6SAW 7218'
    from_place = 'Delhi'
    sdate = datetime.strptime('21-08-2024', '%d-%m-%Y')
    edate = datetime.strptime('20-09-2024', '%d-%m-%Y')
    distance = 62
    odometer_start = 200999
    total_amount_display = 19150
    driver_name = 'suresh'
    price_per_km = 8
    create_log_sheet(name, code, plate, sdate, edate, from_place, distance, price_per_km, total_amount_display,
                     driver_name, odometer_start)


